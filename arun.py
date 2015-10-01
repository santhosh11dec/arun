#import xlrd 
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt

#book = xlrd.open_workbook("/home/santhosh/Downloads/Fatigue crack growth data for SENT speciemn.xlsx")
wb = load_workbook(filename = '/home/santhosh/Downloads/Fatigue crack growth data for SENT speciemn.xlsx')
ws = wb.get_sheet_by_name("Sheet1")
firstRow = 3
zero = 0
out = np.zeros(1000)
X = 0

for limits in range(0,1000):
  try:
    for i in range(firstRow,firstRow+7):
      x = np.array([float(ws['B'+str(firstRow)].value),float(ws['B'+str(firstRow+1)].value),\
      float(ws['B'+str(firstRow+2)].value),float(ws['B'+str(firstRow+3)].value),\
      float(ws['B'+str(firstRow+4)].value),float(ws['B'+str(firstRow+5)].value),\
      float(ws['B'+str(firstRow+6)].value)])
      
      y = np.array([float(ws['C'+str(firstRow)].value),float(ws['C'+str(firstRow+1)].value),\
      float(ws['C'+str(firstRow+2)].value),float(ws['C'+str(firstRow+3)].value),\
      float(ws['C'+str(firstRow+4)].value),float(ws['C'+str(firstRow+5)].value),\
      float(ws['C'+str(firstRow+6)].value)])
      
      z = np.polyfit(x, y, 2)
      diff = np.array([2*z[0],z[1],0])
      output = diff[0]*float(ws['B'+str(firstRow+3)].value) + diff[1]*float(ws['B'+str(firstRow+3)].value)
      print output
      
      out[X] = output
      X = X+1     
      if zero in x: 
        break
      #print x,y
  except IndexError:
      break     
  firstRow = firstRow + 1
#print float(ws['B18'].value)
plt.plot([1,2,3,4]) 
